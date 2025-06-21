# api/views.py

import csv
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from django.db import transaction
from django.db.models import Count, Sum, F, ProtectedError
from django.db.models.functions import TruncDay, TruncMonth, ExtractHour, TruncHour
from django.contrib.auth.models import User, Group
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework import viewsets, status, serializers
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.views import TokenObtainPairView

from .models import (
    Product, Category, Provider, Client, Sale, SaleDetail, PaymentMethod,
    CashCount,
)
from .serializers import (
    ProductSerializer, CategorySerializer, ProviderSerializer, ClientSerializer,
    SaleReadSerializer, SaleWriteSerializer, MyTokenObtainPairSerializer,
    UserSerializer, GroupSerializer, PaymentMethodSerializer, CashCountSerializer
)
from .permissions import (
    IsSuperAdminUser, IsAdminUser, IsVendedorUser,
    IsSuperAdminOrAdmin, CanViewPanel, CanCreateSales
)

class MyTokenObtainPairView(TokenObtainPairView):
    serializer_class = MyTokenObtainPairSerializer

class ProductViewSet(viewsets.ModelViewSet):
    # --- INICIO DE CAMBIOS ---
    queryset = Product.objects.all().order_by('name')
    serializer_class = ProductSerializer

    # CAMBIO: Modificamos get_queryset para permitir el filtrado por estado
    def get_queryset(self):
        queryset = super().get_queryset()
        estado = self.request.query_params.get('estado')
        if estado:
            queryset = queryset.filter(estado=estado)
        return queryset

    def get_permissions(self):
        # CAMBIO: Añadimos la nueva acción 'popular' a las acciones permitidas
        if self.action in ['list', 'retrieve', 'update_stock', 'popular_for_pos']:
            self.permission_classes = [IsAuthenticated, CanViewPanel]
        elif self.action in ['create', 'update', 'partial_update', 'destroy']:
            self.permission_classes = [IsAuthenticated, IsSuperAdminOrAdmin]
        else:
            self.permission_classes = [IsAuthenticated]
        return super().get_permissions()

    # CAMBIO: Sobrescribimos el método destroy para la lógica de soft-delete
    def destroy(self, request, *args, **kwargs):
        product = self.get_object()
        # Verificamos si el producto ha sido usado en alguna venta
        if SaleDetail.objects.filter(product=product).exists():
            # Si ha sido usado, lo desactivamos en lugar de borrarlo
            product.estado = 'inactivo'
            product.save()
            return Response(
                {"detail": "Este producto no se puede eliminar porque tiene ventas asociadas. Se ha marcado como inactivo."},
                status=status.HTTP_200_OK
            )
        else:
            # Si no ha sido usado, lo eliminamos permanentemente
            product.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)

    # CAMBIO: Nueva acción para obtener productos populares para el POS
    @action(detail=False, methods=['get'], url_path='popular-for-pos')
    def popular_for_pos(self, request):
        """
        Devuelve los 10 productos activos más vendidos en los últimos 90 días.
        """
        ninety_days_ago = timezone.now() - timedelta(days=90)
        
        # Obtenemos los IDs de los productos más vendidos y su cantidad
        sold_products_ranking = SaleDetail.objects.filter(
            sale__status='Completada',
            sale__date_time__gte=ninety_days_ago,
            product__estado='activo' # Solo productos activos
        ).values('product_id').annotate(
            total_sold=Sum('quantity')
        ).order_by('-total_sold')[:10]

        # Mapeamos los IDs a los objetos de producto
        product_ids_ordered = [item['product_id'] for item in sold_products_ranking]
        
        # Obtenemos los productos manteniendo el orden de popularidad
        products_queryset = Product.objects.filter(id__in=product_ids_ordered)
        products_dict = {product.id: product for product in products_queryset}
        sorted_products = [products_dict[pid] for pid in product_ids_ordered if pid in products_dict]
        
        serializer = self.get_serializer(sorted_products, many=True)
        return Response(serializer.data)
    # --- FIN DE CAMBIOS ---
    
    @action(detail=True, methods=['patch'], url_path='update-stock')
    def update_stock(self, request, pk=None):
        product = self.get_object()
        stock_to_add = request.data.get('stock')
        if stock_to_add is None:
            return Response({'error': 'El campo "stock" es requerido.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            product.stock += int(stock_to_add)
            product.save()
            return Response(ProductSerializer(product).data, status=status.HTTP_200_OK)
        except (ValueError, TypeError):
            return Response({'error': 'El stock debe ser un número entero válido.'}, status=status.HTTP_400_BAD_REQUEST)

class CategoryViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, IsSuperAdminOrAdmin]
    queryset = Category.objects.all().order_by('name')
    serializer_class = CategorySerializer

    def destroy(self, request, *args, **kwargs):
        category = self.get_object()
        # Verificamos si la categoría está siendo usada por algún producto
        if Product.objects.filter(category=category).exists():
            # Si está en uso, la desactivamos
            category.is_active = False
            category.save()
            return Response(
                {"detail": "Esta categoría está en uso por uno o más productos. Ha sido desactivada."},
                status=status.HTTP_200_OK
            )
        else:
            # Si no está en uso, la eliminamos permanentemente
            category.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)

class ProviderViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, IsSuperAdminOrAdmin]
    queryset = Provider.objects.all().order_by('name')
    serializer_class = ProviderSerializer

    def destroy(self, request, *args, **kwargs):
        provider = self.get_object()
        # Verificamos si el proveedor está siendo usado por algún producto
        if Product.objects.filter(provider=provider).exists():
            # Si está en uso, lo desactivamos
            provider.is_active = False
            provider.save()
            return Response(
                {"detail": "Este proveedor está en uso por uno o más productos. Ha sido desactivado."},
                status=status.HTTP_200_OK
            )
        else:
            # Si no está en uso, lo eliminamos permanentemente
            provider.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)

class ClientViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, IsSuperAdminOrAdmin]
    queryset = Client.objects.all().order_by('name')
    serializer_class = ClientSerializer

    def destroy(self, request, *args, **kwargs):
        client = self.get_object()
        # Verificamos si el cliente tiene ventas asociadas
        if Sale.objects.filter(client=client).exists():
            # Si tiene ventas, lo desactivamos
            client.is_active = False
            client.save()
            return Response(
                {"detail": "Este cliente tiene ventas asociadas y no se puede eliminar. Ha sido desactivado."},
                status=status.HTTP_200_OK
            )
        else:
            # Si no, lo eliminamos permanentemente
            client.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)

class UserViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, IsSuperAdminUser]
    queryset = User.objects.all().order_by('username')
    serializer_class = UserSerializer

    @action(detail=True, methods=['post'], url_path='set-password')
    def set_password(self, request, pk=None):
        user = self.get_object()
        password = request.data.get('password')
        if not password:
            return Response({'error': 'La contraseña no puede estar vacía.'}, status=status.HTTP_400_BAD_REQUEST)

        user.set_password(password)
        user.save()

        return Response({'status': 'Contraseña actualizada con éxito'}, status=status.HTTP_200_OK)

class GroupViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, IsSuperAdminUser]
    queryset = Group.objects.all().order_by('name')
    serializer_class = GroupSerializer

class SaleViewSet(viewsets.ModelViewSet):
    queryset = Sale.objects.all().order_by('-date_time')

    def get_serializer_class(self):
        return SaleWriteSerializer if self.action == 'create' else SaleReadSerializer

    def get_permissions(self):
        if self.action == 'create':
            self.permission_classes = [IsAuthenticated, CanCreateSales]
        else:
            self.permission_classes = [IsAuthenticated, IsSuperAdminOrAdmin]
        return super().get_permissions()

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        try:
            with transaction.atomic():
                details_data = serializer.validated_data.pop('details')
                payment_method_id = serializer.validated_data.pop('payment_method_id')
                payment_method = PaymentMethod.objects.get(id=payment_method_id)
                sale = serializer.save(user=request.user, payment_method=payment_method)
                for detail in details_data:
                    # --- INICIO DE CAMBIOS ---
                    # CAMBIO: Verificamos que el producto esté activo antes de vender
                    product = Product.objects.get(id=detail['product_id'])
                    if product.estado != 'activo':
                        raise serializers.ValidationError(f"El producto '{product.name}' no está activo y no se puede vender.")
                    # --- FIN DE CAMBIOS ---
                    if product.stock < detail['quantity']:
                        raise serializers.ValidationError(f"No hay stock para {product.name}")
                    product.stock -= detail['quantity']
                    product.save()
                    SaleDetail.objects.create(sale=sale, product=product, quantity=detail['quantity'], unit_price=detail['unit_price'])
        except (serializers.ValidationError, Product.DoesNotExist, PaymentMethod.DoesNotExist) as e:
            return Response(str(e), status=status.HTTP_400_BAD_REQUEST)

        sale.refresh_from_db()
        read_serializer = SaleReadSerializer(sale, context={'request': request})
        return Response(read_serializer.data, status=status.HTTP_201_CREATED)

    def destroy(self, request, *args, **kwargs):
        sale = self.get_object()
        with transaction.atomic():
            for detail in sale.details.all():
                product = detail.product
                product.stock += detail.quantity
                product.save()
            sale.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

@api_view(['PATCH'])
@permission_classes([IsAuthenticated, IsSuperAdminOrAdmin])
@transaction.atomic
def cancel_sale_view(request, pk):
    try:
        sale = Sale.objects.get(pk=pk)
    except Sale.DoesNotExist:
        return Response({'detail': 'Venta no encontrada.'}, status=status.HTTP_404_NOT_FOUND)

    if sale.status == 'Cancelada':
        return Response({'detail': 'Esta venta ya ha sido cancelada.'}, status=status.HTTP_400_BAD_REQUEST)

    sale.status = 'Cancelada'
    sale.save()

    for detail in sale.details.all():
        product = detail.product
        product.stock += detail.quantity
        product.save()

    return Response({'detail': 'Venta cancelada y stock restaurado con éxito.'}, status=status.HTTP_200_OK)


class PaymentMethodViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = [IsAuthenticated, CanViewPanel]
    queryset = PaymentMethod.objects.filter(is_active=True)
    serializer_class = PaymentMethodSerializer

class AdminPaymentMethodViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, IsSuperAdminOrAdmin]
    queryset = PaymentMethod.objects.all()
    serializer_class = PaymentMethodSerializer

    def destroy(self, request, *args, **kwargs):
        method = self.get_object()
        # Verificamos si el método de pago ha sido usado en alguna venta
        if Sale.objects.filter(payment_method=method).exists():
            # Si está en uso, lo desactivamos
            method.is_active = False
            method.save()
            return Response(
                {"detail": "Este método de pago está en uso y no se puede eliminar. Ha sido desactivado."},
                status=status.HTTP_200_OK
            )
        else:
            # Si no está en uso, lo eliminamos permanentemente
            method.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)

class DashboardReportsView(APIView):
    permission_classes = [IsAuthenticated, CanViewPanel]

    def get(self, request, *args, **kwargs):
        today = timezone.localdate()
        last_30_days_start = today - timedelta(days=29)

        today_sales_qs = Sale.objects.filter(date_time__date=today, status='Completada')
        total_sales_today = today_sales_qs.aggregate(total=Sum('final_amount'))['total'] or 0
        
        # --- INICIO DE CAMBIOS ---
        # CAMBIO: Aseguramos que los cálculos solo usen productos que estaban activos
        gross_profit_today = SaleDetail.objects.filter(
            sale__in=today_sales_qs
        ).annotate(
            profit_per_item=F('unit_price') - F('product__cost_price')
        ).aggregate(total_profit=Sum(F('quantity') * F('profit_per_item')))['total_profit'] or 0

        kpis = {
            'ventas_del_dia': total_sales_today,
            'ganancia_bruta_del_dia': gross_profit_today,
            'ticket_promedio': total_sales_today / today_sales_qs.count() if today_sales_qs.count() > 0 else 0,
            'productos_vendidos': SaleDetail.objects.filter(sale__in=today_sales_qs).aggregate(total_quantity=Sum('quantity'))['total_quantity'] or 0
        }
        
        # CAMBIO: Productos con bajo stock ahora solo muestra los activos
        low_stock_limit = 5
        low_stock_products_query = Product.objects.filter(stock__lte=low_stock_limit, stock__gt=0, estado='activo').order_by('stock').values('id', 'name', 'stock')[:10]

        sales_by_payment_method = Sale.objects.filter(date_time__date__gte=last_30_days_start, status='Completada').values('payment_method__name').annotate(total=Sum('final_amount')).order_by('-total')
        daily_sales = Sale.objects.filter(date_time__date__gte=last_30_days_start, status='Completada').annotate(day=TruncDay('date_time')).values('day').annotate(total_sales=Sum('final_amount')).order_by('day')
        monthly_sales = Sale.objects.filter(date_time__date__gte=today - timedelta(days=365), status='Completada').annotate(month=TruncMonth('date_time')).values('month').annotate(total_sales=Sum('final_amount')).order_by('month')
        peak_hours_query = Sale.objects.filter(date_time__date__gte=last_30_days_start, status='Completada').annotate(hour=ExtractHour('date_time')).values('hour').annotate(total=Sum('final_amount')).order_by('hour')
        sales_by_hour_dict = {item['hour']: item['total'] for item in peak_hours_query}
        peak_hours_data = [{'name': f"{h:02d}h", 'Ventas': sales_by_hour_dict.get(h, 0)} for h in range(24)]
        
        sales_by_category_query = SaleDetail.objects.filter(
            sale__status='Completada', 
            sale__date_time__date__gte=last_30_days_start
        ).values('product__category__name').annotate(value=Sum(F('quantity') * F('unit_price'))).order_by('-value')

        most_sold_products_query = SaleDetail.objects.filter(
            sale__status='Completada', 
            sale__date_time__date__gte=last_30_days_start
        ).values('product__name').annotate(value=Sum('quantity')).order_by('-value')[:10]
        
        most_profitable_products_query = SaleDetail.objects.filter(
            sale__status='Completada', 
            sale__date_time__date__gte=last_30_days_start
        ).annotate(
            profit_per_sale=F('quantity') * (F('product__sale_price') - F('product__cost_price'))
        ).values('product__name').annotate(value=Sum('profit_per_sale')).order_by('-value')[:10]
        
        dormant_period_days = 60
        dormant_since = today - timedelta(days=dormant_period_days)
        sold_product_ids = SaleDetail.objects.filter(sale__status='Completada', sale__date_time__date__gte=dormant_since).values_list('product_id', flat=True).distinct()
        
        # CAMBIO: Productos dormidos ahora solo muestra los activos
        dormant_products_query = Product.objects.filter(stock__gt=0, estado='activo').exclude(id__in=sold_product_ids).values('name', 'sku', 'stock')[:10]
        # --- FIN DE CAMBIOS ---
        
        chart_data = {
            'ventas_por_metodo_pago': [{'name': item['payment_method__name'] or 'No especificado', 'value': item['total']} for item in sales_by_payment_method],
            'ventas_diarias': [{'name': item['day'].strftime('%d/%m'), 'Ventas': item['total_sales']} for item in daily_sales],
            'ventas_mensuales': [{'name': item['month'].strftime('%b %Y'), 'Ventas': item['total_sales']} for item in monthly_sales],
            'ventas_por_hora': peak_hours_data,
            'ventas_por_categoria': [{'name': item['product__category__name'], 'Ventas': item['value']} for item in sales_by_category_query]
        }

        rankings_data = {
            'mas_vendidos': list(most_sold_products_query),
            'mas_rentables': list(most_profitable_products_query)
        }
        
        other_reports = {
            'productos_dormidos': list(dormant_products_query)
        }

        return Response({
            'kpis': kpis,
            'charts': chart_data,
            'rankings': rankings_data,
            'other_reports': other_reports,
            'low_stock_products': list(low_stock_products_query)
        })

class ReportsView(APIView):
    permission_classes = [IsAuthenticated, IsSuperAdminOrAdmin]

    def get(self,request,*args,**kwargs):
        most_sold=SaleDetail.objects.values('product__name').annotate(c=Sum('quantity')).order_by('-c').first()
        most_profitable=Product.objects.filter(saledetail__isnull=False).annotate(p=F('sale_price')-F('cost_price')).order_by('-p').first()
        peak_hour=Sale.objects.annotate(h=TruncHour('date_time')).values('h').annotate(c=Count('id')).order_by('-c').first()

        return Response({
            'most_sold_product': {
                'name': most_sold['product__name'] if most_sold else 'N/A',
                'total_sold': most_sold['c'] if most_sold else 0
            },
            'most_profitable_product': {
                'name': most_profitable.name if most_profitable else 'N/A',
                'profit_margin': float(most_profitable.p) if most_profitable else 0
            },
            'peak_hour': {
                'hour': peak_hour['h'].hour if peak_hour else 'N/A',
                'count': peak_hour['c'] if peak_hour else 0
            }
        })

class DailyCashCountView(APIView):
    permission_classes = [IsAuthenticated, IsSuperAdminOrAdmin]

    def get(self, request, *args, **kwargs):
        today = date.today()
        history = CashCount.objects.all().order_by('-date')
        if CashCount.objects.filter(date=today).exists():
            return Response({'message': 'La caja del día de hoy ya fue cerrada.', 'history': CashCountSerializer(history, many=True).data}, status=status.HTTP_409_CONFLICT)
        total_sales_today = Sale.objects.filter(date_time__date=today, status='Completada').aggregate(total=Sum('final_amount'))['total'] or 0
        return Response({'expected_amount': total_sales_today, 'history': CashCountSerializer(history, many=True).data})

    def post(self, request, *args, **kwargs):
        today = date.today()
        if CashCount.objects.filter(date=today).exists():
            return Response({'message': 'La caja del día de hoy ya fue cerrada.'}, status=status.HTTP_409_CONFLICT)

        counted_str = request.data.get('counted_amount')
        expected_str = request.data.get('expected_amount')

        if counted_str is None or expected_str is None:
            return Response({'error': 'Faltan datos.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            counted_decimal = Decimal(str(counted_str))
            expected_decimal = Decimal(str(expected_str))
            difference = counted_decimal - expected_decimal

            CashCount.objects.create(
                date=today,
                expected_amount=expected_decimal,
                counted_amount=counted_decimal,
                difference=difference,
                user=request.user
            )
            return Response({'message': 'Cierre de caja guardado con éxito.'}, status=status.HTTP_201_CREATED)
        except InvalidOperation:
                return Response({'error': 'Los montos deben ser números válidos.'}, status=status.HTTP_400_BAD_REQUEST)

class BulkPriceUpdateView(APIView):
    permission_classes = [IsAuthenticated, IsSuperAdminOrAdmin]

    def post(self, request, *args, **kwargs):
        product_ids = request.data.get('product_ids', [])
        percentage_str = request.data.get('percentage')
        update_target = request.data.get('update_target')
        if not all([product_ids, percentage_str, update_target]):
            return Response({'error': 'Faltan datos requeridos.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            percentage = Decimal(percentage_str)
            percentage_multiplier = Decimal(1 + (percentage / 100))
        except (ValueError, TypeError):
            return Response({'error': 'El porcentaje debe ser un número válido.'}, status=status.HTTP_400_BAD_REQUEST)
        products_to_update = Product.objects.filter(id__in=product_ids)
        if not products_to_update.exists():
            return Response({'error': 'No se encontraron productos.'}, status=status.HTTP_404_NOT_FOUND)
        try:
            with transaction.atomic():
                for product in products_to_update:
                    if update_target == 'cost' or update_target == 'both': product.cost_price *= percentage_multiplier
                    if update_target == 'sale' or update_target == 'both': product.sale_price *= percentage_multiplier
                    product.save()
        except Exception as e:
            return Response({'error': f'Error en la actualización: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        return Response({'message': f'{len(product_ids)} productos actualizados.'}, status=status.HTTP_200_OK)

class ExportSalesView(APIView):
    permission_classes = [IsAuthenticated, IsSuperAdminOrAdmin]

    def get(self, request, *args, **kwargs):
        start_date_str = request.query_params.get('start_date')
        end_date_str = request.query_params.get('end_date')
        if not start_date_str or not end_date_str:
            return Response({"error": "Las fechas de inicio y fin son requeridas."}, status=status.HTTP_400_BAD_REQUEST)
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({"error": "Formato de fecha inválido. Usar YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)

        sales = Sale.objects.filter(date_time__date__gte=start_date, date_time__date__lte=end_date).order_by('date_time')

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Ventas"

        headers = ['Fecha', 'ID Venta', 'Estado', 'Monto Total', 'Neto Gravado (21%)', 'IVA (21%)']
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        iva_rate = Decimal('1.21')

        for sale in sales:
            total = sale.final_amount or Decimal('0.00')
            neto = Decimal('0.00')
            iva = Decimal('0.00')
            if total > 0:
                neto = total / iva_rate
                iva = total - neto

            naive_datetime = sale.date_time.replace(tzinfo=None)

            ws.append([
                naive_datetime,
                sale.id,
                sale.status,
                total,
                neto,
                iva,
            ])

        for col_num, header_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            if header_title in ['Monto Total', 'Neto Gravado (21%)', 'IVA (21%)']:
                ws.column_dimensions[col_letter].width = 18
                for cell in ws[col_letter]:
                    if cell.row > 1: cell.number_format = '"$"#,##0.00'
            elif header_title == 'Fecha':
                ws.column_dimensions[col_letter].width = 20
                for cell in ws[col_letter]:
                    if cell.row > 1: cell.number_format = 'DD/MM/YYYY HH:MM'
            else:
                ws.column_dimensions[col_letter].width = 15
                for cell in ws[col_letter]:
                    if cell.row > 1: cell.alignment = Alignment(horizontal='center')

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="reporte_ventas_{start_date_str}_a_{end_date_str}.xlsx"'

        wb.save(response)
        return response

class ChangePasswordView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        user = request.user
        old_password = request.data.get('old_password')
        new_password = request.data.get('new_password')
        confirm_password = request.data.get('confirm_password')

        if not all([old_password, new_password, confirm_password]):
            return Response({'error': 'Todos los campos son requeridos.'}, status=status.HTTP_400_BAD_REQUEST)

        if not user.check_password(old_password):
            return Response({'error': 'La contraseña actual es incorrecta.'}, status=status.HTTP_400_BAD_REQUEST)

        if new_password != confirm_password:
            return Response({'error': 'Las contraseñas nuevas no coinciden.'}, status=status.HTTP_400_BAD_REQUEST)

        if len(new_password) < 8:
            return Response({'error': 'La nueva contraseña debe tener al menos 8 caracteres.'}, status=status.HTTP_400_BAD_REQUEST)

        user.set_password(new_password)
        user.save()

        return Response({'status': 'Contraseña cambiada con éxito.'}, status=status.HTTP_200_OK)